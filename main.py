import os
import csv
import base64
import imghdr
import re
from flask import Flask, render_template, request, jsonify
from PyPDF2 import PdfReader
import docx
import openpyxl
import chardet
from difflib import SequenceMatcher

app = Flask(__name__)
DATA_DIR = 'data'

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)


# קריאת קובץ docx
def read_docx(file_path):
    if os.path.basename(file_path).startswith('~$'):
        return "", []

    doc = docx.Document(file_path)
    content = ''
    images = []

    for para in doc.paragraphs:
        content += para.text + '\n'

    for rel in doc.part.rels.values():
        if "image" in rel.target_ref:
            image_data = rel.target_part.blob
            image_type = imghdr.what(None, h=image_data) or 'jpeg'
            img_base64 = base64.b64encode(image_data).decode('utf-8')
            images.append(f"data:image/{image_type};base64,{img_base64}")

    return content, images


# קריאת קובץ csv
def read_csv(file_path):
    content = ''
    with open(file_path, 'rb') as file:
        raw_data = file.read(10000)
        result = chardet.detect(raw_data)
        encoding = result['encoding']

    try:
        with open(file_path, 'r', encoding=encoding) as file:
            reader = csv.reader(file)
            for row in reader:
                content += ' | '.join(row) + '\n'
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='ISO-8859-1') as file:
            reader = csv.reader(file)
            for row in reader:
                content += ' | '.join(row) + '\n'

    return content


# קריאת PDF
def read_pdf(file_path):
    with open(file_path, 'rb') as file:
        reader = PdfReader(file)
        content = ''
        for page in reader.pages:
            content += page.extract_text()
        return content


# קריאת Excel
def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    content = ''
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                content += str(cell.value) + ' '
    return content


# ניקוי מחרוזות
def preprocess_question(question):
    question = question.lower()
    question = re.sub(r'[^\w\s]', '', question)
    question = re.sub(r'\s+', ' ', question).strip()
    return question


# חילוץ מילים מטקסט (עבור זיהוי מדויק יותר במילים עם תווים מסביב)
def get_words(text):
    # מחזיר רשימת מילים בלבד, ללא תווים מיוחדים
    return re.findall(r'\b\w+\b', text)


# פונקציה לבדוק אם מילה מכילה תוספת של תו אחד בלבד או יותר
def keyword_similar_or_surrounded(keyword, word, threshold=0.9):
    # תנאי מיוחד – "מי" ו־"מה" נחשבות שוות ל־"מידע"
    if ((keyword in ['מי', 'מה', 'תן'] and word == 'מידע')
            or (word in ['מי', 'מה', 'תן'] and keyword == 'מידע')):
        return True

    if len(keyword) < 3 or len(word) < 3:
        return False
    if keyword in word or word in keyword:
        return True
    # מאפשר תו אחד לפני או אחרי המילה
    if re.fullmatch(rf'.{{0,1}}{re.escape(keyword)}.{{0,1}}', word):
        return True
    return SequenceMatcher(None, keyword, word).ratio() >= threshold


# חיפוש קובץ מתאים
def find_best_match(question):
    files = os.listdir(DATA_DIR)
    question_clean = preprocess_question(question)
    keywords = question_clean.split()

    best_file = None
    best_filename = ""
    highest_match_count = 0

    for filename in files:
        full_path = os.path.join(DATA_DIR, filename)
        if os.path.isfile(full_path):
            filename_base = preprocess_question(os.path.splitext(filename)[0])
            filename_words = set(filename_base.split())

            match_count = 0
            for keyword in keywords:
                for word in filename_words:
                    if keyword_similar_or_surrounded(keyword, word):
                        match_count += 1

            if match_count > highest_match_count:
                highest_match_count = match_count
                best_file = full_path
                best_filename = filename

    if not best_file:
        return None

    content = ''
    images = []
    if best_file.endswith('.txt'):
        with open(best_file, 'r', encoding='utf-8') as f:
            content = f.read()
    elif best_file.endswith('.csv'):
        content = read_csv(best_file)
    elif best_file.endswith('.pdf'):
        content = read_pdf(best_file)
    elif best_file.endswith('.docx'):
        content, images = read_docx(best_file)
    elif best_file.endswith('.xlsx'):
        content = read_excel(best_file)

    # אם זה קובץ Word (docx) – להחזיר תמיד את התוכן, גם אם אין התאמה
    if best_file.endswith('.docx'):
        return {
            'filename': best_filename,
            'content': content,
            'images': images
        }

    # ניתוח לפי SECTION אם מדובר בקובץ HTML-דמוי טקסט
    sections = re.findall(r'<section.*?>(.*?)</section>', content,
                          re.DOTALL | re.IGNORECASE)
    matched_sections = []
    question_keywords = set(keywords)

    for section in sections:
        section_clean = preprocess_question(section)

        # בדיקת התאמה ל־<h1>
        h1_match = False
        h1_matches = re.findall(r'<h1.*?>(.*?)</h1>', section,
                                re.DOTALL | re.IGNORECASE)
        for h1 in h1_matches:
            h1_clean = preprocess_question(h1)
            h1_words = get_words(h1_clean)
            for keyword in question_keywords:
                if any(
                        keyword_similar_or_surrounded(keyword, w)
                        for w in h1_words):
                    h1_match = True
                    break
            if h1_match:
                break

        # בדיקת התאמה להערות <!-- -->
        comment_match = False
        comments = re.findall(r'<!--(.*?)-->', section, re.DOTALL)
        for comment in comments:
            comment_clean = preprocess_question(comment)
            comment_words = get_words(comment_clean)
            for keyword in question_keywords:
                if any(
                        keyword_similar_or_surrounded(keyword, w)
                        for w in comment_words):
                    comment_match = True
                    break
            if comment_match:
                break

        # בדיקת strong
        strong_lines = []
        lines = section.splitlines()
        for line in lines:
            strong_words = re.findall(r'<strong>(.*?)</strong>', line,
                                      re.IGNORECASE)
            strong_words_clean = []
            for s in strong_words:
                s_clean = preprocess_question(s)
                strong_words_clean.extend(get_words(s_clean))
            if any(
                    keyword_similar_or_surrounded(keyword, word)
                    for keyword in question_keywords
                    for word in strong_words_clean):
                strong_lines.append(line.strip())

        # לוגיקה לפי הכללים:
        if (h1_match or comment_match) and strong_lines:
            matched_sections.extend(strong_lines)
        elif h1_match or comment_match:
            matched_sections.append(section.strip())

    # בדיקה אם בשאלה יש את המילה "כל"
    return_all = 'כל' in question_clean.split()

    if matched_sections:
        if return_all:
            final_content = '\n\n'.join(matched_sections)
        else:
            final_content = matched_sections[0]
    else:
        final_content = 'אין מידע על כך.'

    return {
        'filename': best_filename,
        'content': final_content,
        'images': images
    }


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/ask', methods=['POST'])
def ask():
    question = request.form['question']
    best_match = find_best_match(question)

    if best_match:
        return jsonify({
            'answer': best_match['content'],
            'filename': best_match['filename'],
            'images': best_match.get('images', [])
        })
    else:
        return jsonify({'answer': 'מה השאלה?.', 'images': []})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3000, debug=True)
